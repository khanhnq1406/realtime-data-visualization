from multiprocessing.spawn import import_main_path
from time import sleep
import cv2
from numpy import product
from pyzbar import pyzbar
from openpyxl import Workbook,load_workbook
from datetime import datetime
import difflib
from snap7.util import*
from snap7.types import*
import database
from database import items
import connect_plc
from connect_plc import client,UNIT
import connect_plc_s7
from connect_plc_s7 import product_name1,product_name2,product_name3,product_status,plc,DB_NUMBER,START_ADDRESS,SIZE
# def read_barcodes(frame):
#     barcodes = pyzbar.decode(frame)
#     for barcode in barcodes:
#         x,y,w,h = barcode.rect
#         barcode_text=barcode.data.decode('utf-8')
#         print (barcode_text)
#         cv2.rectangle(frame,(x,y),(x+w,y+h),(0,255,0),2)
#         cv2.imshow('Barcode reader',frame)
#         return frame

def excel(barcode_text):
    
    now = datetime.now()
    dtString = now.strftime('%m/%d/%Y, %H:%M:%S')
    wb = load_workbook('E:/test/barcode.xlsx')
    ws=wb.active
    for count in range(1, 100):
        name = 'A' + str(count)
        time='B' +str(count)
        if ws[name].value is None:
            ws[name] = barcode_text
            ws[time]=dtString
            wb.save('E:/test/barcode.xlsx')
            wb.close()
            break
        else:
            pass
camera = cv2.VideoCapture(1)
#HCM = 1, KH = 2, DN = 3
xylanh=['','','']
xylanh[0]= product_name1
xylanh[1]= product_name2
xylanh[2]= product_name3

while True:
    # ret,frame=camera.read()
    # frame=read_barcodes(frame)
    # cv2.imshow('Barcode reader',frame)
    # if cv2.waitKey(1) & 0xFF == 27:
    #    break
    r, frame = camera.read() 
    # cv2.imshow('frame', frame)
    barcodes = pyzbar.decode(frame)
    db = plc.db_read(DB_NUMBER, START_ADDRESS, SIZE)
    product_status = bool(db[768])
    # print(product_status)
    if product_status == True:   
        db = plc.db_read(DB_NUMBER, START_ADDRESS, SIZE)
        product_name1 = db[2:256].decode('UTF-8').strip('\x00')
        product_name2 = db[258:512].decode('UTF-8').strip('\x00')
        product_name3 = db[514:768].decode('UTF-8').strip('\x00')
        print(product_name1)
        print(product_name2)
        print(product_name3)
        xylanh[0]= product_name1
        xylanh[1]= product_name2
        xylanh[2]= product_name3
    for barcode in barcodes:
        x,y,w,h = barcode.rect
        barcode_text=barcode.data.decode('utf-8')
        excel(barcode_text)
        
        # print (barcode_text)
        # print (barcode.type)
        for item in items:           
            if barcode_text==item[1]:
                for i in range(len(xylanh)):
                    output = str(int(difflib.SequenceMatcher(None, xylanh[i].strip().lower(), item[2].strip().lower()).ratio()*100))
                    if int(output)>70:
                        print("Day xylanh"+str(i+1))
                        client.write_register(0,i+1,unit=UNIT)
                        print(item[2])
        sleep(1)
        
        # cv2.rectangle(frame,(x,y),(x+w,y+h),(0,255,0),2)
    cv2.imshow('Barcode reader',frame)
    if cv2.waitKey(1) & 0xFF == ord('q'): 
        break
camera.release()
cv2.destroyAllWindows()